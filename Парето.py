import openpyxl
from itertools import combinations
names = ['Двигатели.xlsx','Колёсные установки.xlsx','Редукторы.xlsx','Тормоза.xlsx','Муфты двигатель-редуктор.xlsx','Муфты редуктор-колесо.xlsx']
parametr = ['P,кВт','P,кН','T,Н*м','T,Н*м','T,Н*м','T,Н*м']
extra = []
speed =  float(input('Введите необходимую скорость передвижения крана v,м/с '))
for i in range(len(names)):
    print('\n' + names[i])
    print('Введите минимальное и максимальное значения для параметра ' + parametr[i])
    Min = float(input('Min = '))
    Max = float(input('Max = '))
    extra.append([Min,Max])

def func1(run,Min,Max):
    sheet = run.active
    qwerty = []
    result = {}
    if sheet.title == 'Редукторы':
        par = float('inf')
        for i in range(2,sheet.max_row+1):
            if par > sheet['G' + str(i)].value:
                par = sheet['G' + str(i)].value
    else:
        par = 0
        try:
            for i in range(2,sheet.max_row+1):
                if par < sheet['G' + str(i)].value:
                    par = sheet['G' + str(i)].value
        except:
            par = 1              
    for i in range(2,sheet.max_row+1):
        if i in qwerty or sheet['F'+str(i)].value < Min or sheet['F'+str(i)].value > Max:
            continue
        old_row = [sheet['C'+str(i)].value,sheet['D'+str(i)].value,sheet['E'+str(i)].value]
        for j in range(i+1,sheet.max_row+1):
            new_row = [sheet['C'+str(j)].value,sheet['D'+str(j)].value,sheet['E'+str(j)].value]
            a1,b1,c1 = old_row
            a2,b2,c2 = new_row
            if a1<=a2 and b1>=b2 and c1>=c2:
                qwerty.append(i)
            elif a1>=a2 and b1<=b2 and c1<=c2:
                qwerty.append(j)
        if i not in qwerty:
            key = sheet['B1'].value + ' ' + sheet['B'+str(i)].value
            value = list(old_row)
            if sheet.title == 'Редукторы':
                value.append(sheet['G' + str(i)].value/par)
            else:
                try:
                    value.append(par/sheet['G' + str(i)].value)
                except:
                    value.append(par)
            result[key] = value
    return par, result

Dict = {}
mult = 3.14/(speed*60*1000)
for i in range(len(names)):
    wb = openpyxl.load_workbook(names[i])
    par, res = func1(wb,extra[i][0],extra[i][1])
    Dict[names[i]] = res
    if names[i] == 'Редукторы.xlsx':
        par = 1/par
    mult = mult*par

def func2(dict1,dict2):
    Help = dict(list(dict1.items())+list(dict2.items()))
    data = list(combinations(Help,2))
    a = dict1.keys()
    b = dict2.keys()
    KEYS = []
    for lie in data:
        if lie[0] in a and lie[1] in a or lie[0] in b and lie[1] in b :
            continue
        else:
            KEYS.append(lie)
    qwerty = []
    result = {}
    for i in range(len(KEYS)):
        if i in qwerty:
            continue
        a111,b121,c131,d141 = dict1[KEYS[i][0]]
        a211,b221,c231,d241 = dict2[KEYS[i][1]]
        for j in range(i+1,len(KEYS)):
            a112,b122,c132,d142 = dict1[KEYS[j][0]]
            a212,b222,c232,d242 = dict2[KEYS[j][1]]
            if a111+a211<=a112+a212 and b121+b221>=b122+b222 and c131+c231>=c132+c232:
                qwerty.append(i)
            elif a111+a211>=a112+a212 and b121+b221<=b122+b222 and c131+c231<=c132+c232:
                qwerty.append(j)
        if i not in qwerty:
            if ':' not in KEYS[i][0]:
                key1 = KEYS[i][0] + ': ' + ' '.join(map(str,[a111,b121,c131])) + '\n'
            else:
                key1 = KEYS[i][0] + '\n'  
            if ':' not in KEYS[i][1]:
                key2 = KEYS[i][1] + ': ' + ' '.join(map(str,[a211,b221,c231]))
            else:
                key2 = KEYS[i][1]
            key = key1 + key2
            result[key] = [a111+a211,b121+b221,c131+c231,d141*d241]
    return result

b = func2(Dict[names[0]],Dict[names[1]])
b = func2(b,Dict[names[2]])
for k, v in list(b.items()):
    if (mult - v[-1])/mult < 0 or (mult - v[-1])/mult >=0.15:
        del b[k]

for i in range(3,len(names)):
    b = func2(b,Dict[names[i]])
    
number = 1
file = open ('Результаты.txt', 'w')
for k,v in sorted(b.items(),key = lambda x: x[1][0]/(x[1][1]*x[1][2]), reverse = True):
    file.write('\n\nВариант ' + str(number) +':' + '\n')
    file.write(k)
    number += 1
    if number > 5:
        file.close()
        break

    

    
    


    
